from extract_id import *
from get_answers import *
from paper_extraction import *
import sys
import tkinter as tk
from tkinter import filedialog
import openpyxl
import os

def generate_excel_sheet(student_data, output_path="results/bubble_sheet.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Score'
    sheet['C1'] = 'Total'

    for row_num, (student_id, score, fullmark) in enumerate(student_data, start=2):
        sheet.cell(row=row_num, column=1, value=student_id)
        sheet.cell(row=row_num, column=2, value=score)
        sheet.cell(row=row_num, column=3, value=fullmark)

    workbook.save(output_path)

def generate_excel_sheet_answers(student_data, output_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Q'
    sheet['B1'] = 'Answer'
    sheet['C1'] = 'Correct'

    for row_num, (Q, answer, correct) in enumerate(student_data, start=2):
        sheet.cell(row=row_num, column=1, value=Q)
        sheet.cell(row=row_num, column=2, value=answer)
        sheet.cell(row=row_num, column=3, value=correct)

    workbook.save(output_path)


def correct_paper(model_answers_path = "./samples/model_test2.jPG", student_answers_path = "./samples/student_test2.jPG"):

    #get the correct answers
    img_BGR = cv2.imread(model_answers_path, cv2.IMREAD_COLOR)
    img_BGR = cv2.resize(img_BGR, (1200, 1600))
    colored, grayed = extract_paper_region(img_BGR)
    answers_section_gray = grayed[grayed.shape[0]//3 : (grayed.shape[0]-grayed.shape[0]//10), grayed.shape[1]//10 : (grayed.shape[1]-grayed.shape[1]//10)]
    answers_section_colored = colored[colored.shape[0]//3 : (colored.shape[0]-colored.shape[0]//10), colored.shape[1]//10 : (colored.shape[1]-colored.shape[1]//10)]

    res = get_answers(answers_section_colored, answers_section_gray)
    if res is None:
        print("An Error happened while extracting the correct answers from the model answer!")
        sys.exit(1)
    print("answers have been extracted successfully from the model answer..")

    [colored[colored.shape[0]//3 : (colored.shape[0]-colored.shape[0]//10), colored.shape[1]//10 : (colored.shape[1]-colored.shape[1]//10)]
    , model_answers_partitioned] = res

    img_BGR = cv2.imread(student_answers_path, cv2.IMREAD_COLOR)
    img_BGR = cv2.resize(img_BGR, (1200, 1600))

    colored, grayed = extract_paper_region(img_BGR)
    answers_section_gray = grayed[grayed.shape[0]//3 : (grayed.shape[0]-grayed.shape[0]//10), grayed.shape[1]//10 : (grayed.shape[1]-grayed.shape[1]//10)]
    answers_section_colored = colored[colored.shape[0]//3 : (colored.shape[0]-colored.shape[0]//10), colored.shape[1]//10 : (colored.shape[1]-colored.shape[1]//10)]

    id_section_gray = grayed[0 : grayed.shape[0]//3, grayed.shape[1]//10 : (grayed.shape[1]-5*grayed.shape[1]//12)]
    id_section_colored = colored[0 : grayed.shape[0]//3, grayed.shape[1]//10 : (grayed.shape[1]-5*grayed.shape[1]//12)]


    id = extract_id(id_section_colored, id_section_gray)

    if id is None:
        print("An Error happened while extracting the student ID!")
        sys.exit(1)
    print("Current ID = " + id + " is being processed..")

    res = get_answers(answers_section_colored, answers_section_gray, model_answers_partitioned)
    if res is None:
        print("An Error happened while extracting the answers of the student!")
        sys.exit(1)

    [colored[colored.shape[0]//3 : (colored.shape[0]-colored.shape[0]//10), colored.shape[1]//10 : (colored.shape[1]-colored.shape[1]//10)]
    , studnent_answers_partitioned] = res

    correct = 0
    total = 0
    answers = []
    for i in range(len(studnent_answers_partitioned)):
        for j in range(len(studnent_answers_partitioned[i])):
            total+=1
            correct+=studnent_answers_partitioned[i][j] ==  model_answers_partitioned[i][j]
            answers.append((total, chr(ord('A') + studnent_answers_partitioned[i][j]), chr(ord('A') + model_answers_partitioned[i][j])))

    student_image = colored

    print("Correcting of the paper is done successfully!")
    return (id, correct, total, student_image, answers)

def upload_model_answers():
    model_answers_path = filedialog.askopenfilename(title="Select Model Answers", filetypes=[("Image files", "*.jpg;*.jpeg;*.png;*.tif")])
    model_answers_entry.delete(0, tk.END)
    model_answers_entry.insert(0, model_answers_path)

def upload_student_answers():
    student_answers_paths = filedialog.askopenfilenames(title="Select Student Answers", filetypes=[("Image files", "*.jpg;*.jpeg;*.png;*.tif")])
    student_answers_entry.delete(0, tk.END)
    student_answers_entry.insert(0, ', '.join(student_answers_paths))

def compare_answers():
    model_answers_path = model_answers_entry.get()
    student_answers_paths = student_answers_entry.get().split(', ')
    students_data = []
    for answer in student_answers_paths:
        id, correct, total, student_image, answers = correct_paper(model_answers_path, answer)
        students_data.append([id, correct, total])

        if not os.path.exists('results'):
            os.makedirs('results')

        student_folder = os.path.join('results', str(id))
        if not os.path.exists(student_folder):
            os.makedirs(student_folder)

        image_name = f"corrected paper.jpg"
        image_path = os.path.join(student_folder, image_name)
        cv2.imwrite(image_path, student_image)
        
        generate_excel_sheet_answers(answers, "results/"+str(id)+"/answers.xlsx")
        print("The excel sheet for the student is generated successfully!")

    print("The excel sheet for all the students is generated successfully!!")
    generate_excel_sheet(students_data)


root = tk.Tk()
root.title("Answer Comparison Tool")

# Create and set up GUI components
model_answers_label = tk.Label(root, text="Model Answers:")
model_answers_label.grid(row=0, column=0)

model_answers_entry = tk.Entry(root, width=50)
model_answers_entry.grid(row=0, column=1)

upload_model_button = tk.Button(root, text="Upload Model Answers", command=upload_model_answers)
upload_model_button.grid(row=0, column=2)

student_answers_label = tk.Label(root, text="Student Answers:")
student_answers_label.grid(row=1, column=0)

student_answers_entry = tk.Entry(root, width=50)
student_answers_entry.grid(row=1, column=1)

upload_student_button = tk.Button(root, text="Upload Student Answers", command=upload_student_answers)
upload_student_button.grid(row=1, column=2)

compare_button = tk.Button(root, text="Compare Answers", command=compare_answers)
compare_button.grid(row=2, column=1)

result_label = tk.Label(root, text="")
result_label.grid(row=3, column=1)

# Start the Tkinter main loop
root.mainloop()
