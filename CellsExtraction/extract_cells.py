import cv2
import numpy as np
from functools import cmp_to_key
# from PaperExtraction.utils import *
# from Models.test_model import *

import joblib
import cv2
import openpyxl
from openpyxl.styles import PatternFill

svm_model_filename_digit = "svm_model_digits.joblib"
svm_model_filename_symbol = "svm_model_symbols.joblib"
loaded_svm_model_digit = joblib.load(svm_model_filename_digit)
loaded_svm_model_symbol = joblib.load(svm_model_filename_symbol)


target_img_size = (32, 32)
def extract_hog_features(img):
    if img is None or img.size == 0:
        return None
    img = cv2.resize(img, dsize=target_img_size)
    win_size = (32, 32)
    cell_size = (4, 4)
    block_size_in_cells = (2, 2)
    
    block_size = (block_size_in_cells[1] * cell_size[1], block_size_in_cells[0] * cell_size[0])
    block_stride = (cell_size[1], cell_size[0])
    nbins = 9
    hog = cv2.HOGDescriptor(win_size, block_size, block_stride, cell_size, nbins)
    h = hog.compute(img)
    h = h.flatten()
    return h.flatten()


def predict_digit(img):
    hog_features = extract_hog_features(img)
    if hog_features is None:
        return ""
    hog_features = hog_features.reshape(1, -1)

    prediction = loaded_svm_model_digit.predict(hog_features)
    return prediction[0]

def predict_symbol(img):

    hog_features = extract_hog_features(img)
    hog_features = hog_features.reshape(1, -1)

    prediction = loaded_svm_model_symbol.predict(hog_features)
    return prediction[0]



def segement(img):
    kernel = np.ones((5, 5), np.uint8)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 50, 200, 255)
    
    contours, hierarchy = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=lambda ctr: cv2.boundingRect(ctr))
    white_img_large_contours = np.ones(img.shape)
    dimensions_contours = []
    for contour in contours:
        (x, y, w, h) = cv2.boundingRect(contour)
        if(w*h > 50):
            dimensions_contours.append((x, y, w, h))
            cv2.rectangle(white_img_large_contours,(x, y), (x+w, y+h), (0, 0, 255), 3)
    return dimensions_contours, img


def getIdFromImage(img):
    img = cv2.resize(img, (128, 64))
    segmented_dimensions, filtered_img = segement(img)
    cropped_digits = []
    i = 0
    for dimension in segmented_dimensions:
        (x, y, w, h) = dimension
        cropped_digits.append(filtered_img[y-1:y+h+1, x-1:x+w+1])
        i += 1

    predictions = []

    for img in cropped_digits:
        predictions.append(predict_digit(img))

    predictedNumber = ""
    for number in predictions:
        predictedNumber += str(number)
    return predictedNumber

def cmp(a, b):
    br_a = cv2.boundingRect(a)
    br_b = cv2.boundingRect(b)
    if abs(br_a[0] - br_b[0]) <= 5:
        return br_a[1] - br_b[1]
    return br_a[0] - br_b[0]


def sorted_counter(contours):
    return sorted(contours, key=cmp_to_key(cmp))

def string_to_value(s):
    if s == 'box':
        return 0
    elif s == 'correct':
        return 5
    elif s == 'empty':
        return -1
    elif s.startswith('horizontal') and s[10:].isdigit():
        return 5 - int(s[10:])
    elif s.startswith('vertical') and s[8:].isdigit():
        return int(s[8:])
    elif s == 'question':
        return -2
    else:
        # Default case, you can choose to return a specific value or raise an exception
        return None


def extract_cells(img, grid_points_img):
        # Convert grid_points_img to grayscale
    data_to_write = [["Code", "1", "2", "3"]]
    grid_points_gray = cv2.cvtColor(grid_points_img, cv2.COLOR_BGR2GRAY)

    # Threshold the grayscale image (optional, based on your needs)
    _, thresh = cv2.threshold(grid_points_gray, 128, 255, cv2.THRESH_BINARY)

    # Find contours in the thresholded image
    contours, _ = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # Sort contours using your custom comparison function
    contours = sorted_counter(contours)

    rows = []
    for c in range(len(contours)-1):
        x1, y1, w1, h1 = cv2.boundingRect(contours[c])
        x2, y2, w2, h2 = cv2.boundingRect(contours[c+1])
        rows.append(y1)
        if (x1 != x2):
            break

    N = len(rows)
    M = len(contours) // N

    for x in range(0, N-1):
        student_data = []
        for y in range(0, M - 1):
            # Returns the location and width,height for every contour
            if y == 1 or y == 2:
                continue

            x1, y1, w1, h1 = cv2.boundingRect(contours[x + y * N])
            x2, y2, w2, h2 = cv2.boundingRect(contours[x + y * N + 1])
            x3, y3, w3, h3 = cv2.boundingRect(contours[x + (y + 1) * N + 1])
            # crop image
            cell = img[y1 + h1: y3, x2 + w2: x3]
    
            if y == 0:
                student_data.append(getIdFromImage(cell))
            elif y == 3:
                student_data.append(predict_digit(cell))
            else:
                student_data.append(string_to_value(predict_symbol(cell)))
        
        data_to_write.append(student_data)

    generate_excel_sheet(data_to_write)

def generate_excel_sheet(data, file_path='output.xlsx'):
    # Create a new Workbook
    workbook = openpyxl.Workbook()

    # Get the active sheet (the default sheet when a workbook is created)
    sheet = workbook.active

    # Iterate through the data and write it into the cells
    for i, row_data in enumerate(data):
        for j, value in enumerate(row_data):
            # Excel indices start from 1, so we add 1 to i and j
            if value == -1:
                continue
            elif value == -2:
                cell = sheet.cell(row=i + 1, column=j + 1, value="")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
            else:
                sheet.cell(row=i + 1, column=j + 1, value=value)

    # Save the workbook to a file
    workbook.save(file_path)


if __name__ == '__main__':
    img = cv2.imread('CellsExtraction/img.jpg')
    grid_points_img = cv2.imread('CellsExtraction/grid_points_img.jpg')
    extract_cells(img=img, grid_points_img=grid_points_img)


# y = 0 => id
# y = 1 => ignore
# y = 2 => ignore
# y = 3 => digit | question mark | empty cell
# y = 4-5 => symbol | empty cell