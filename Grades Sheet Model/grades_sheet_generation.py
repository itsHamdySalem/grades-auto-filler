import os
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO


def generate_excel_sheet(data=[], file_path='Grades Sheet Model/output.xlsx'):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for i, row_data in enumerate(data):
        for j, value in enumerate(row_data):
            if value == -1:
                continue
            elif value == -2:
                cell = sheet.cell(row=i + 1, column=j + 1, value="")
                cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type='solid')
            else:
                sheet.cell(row=i + 1, column=j + 1, value=value)

    workbook.save(file_path)

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
